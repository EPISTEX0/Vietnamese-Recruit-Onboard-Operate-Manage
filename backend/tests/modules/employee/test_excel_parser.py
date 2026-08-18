"""Unit tests for the Excel parser module."""

import logging
import zipfile
from datetime import date, datetime
from io import BytesIO

import pytest
from openpyxl import Workbook

from src.modules.employee.infrastructure.excel_parser import (
    _MSG_CORRUPTED_FILE,
    _MSG_WRONG_FORMAT,
    parse_excel,
)


def _create_excel(headers: list[str], rows: list[list]) -> bytes:
    """Helper to create an in-memory .xlsx file from headers and rows."""
    wb = Workbook()
    ws = wb.active
    ws.append(headers)
    for row in rows:
        ws.append(row)
    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


class TestParseExcelValidRows:
    """Tests for successfully parsed rows."""

    def test_basic_valid_row(self):
        """A row with all required fields should parse successfully."""
        headers = ["full_name", "email", "phone"]
        rows = [["Nguyen Van A", "a@example.com", "0901234567"]]
        file_bytes = _create_excel(headers, rows)

        parsed, errors = parse_excel(file_bytes)

        assert len(parsed) == 1
        assert len(errors) == 0
        assert parsed[0]["full_name"] == "Nguyen Van A"
        assert parsed[0]["email"] == "a@example.com"
        assert parsed[0]["phone"] == "0901234567"

    def test_multiple_valid_rows(self):
        """Multiple valid rows should all be parsed."""
        headers = ["full_name", "email"]
        rows = [
            ["Nguyen Van A", "a@example.com"],
            ["Tran Thi B", "b@example.com"],
            ["Le Van C", "c@example.com"],
        ]
        file_bytes = _create_excel(headers, rows)

        parsed, errors = parse_excel(file_bytes)

        assert len(parsed) == 3
        assert len(errors) == 0

    def test_all_known_fields(self):
        """All known fields should be parsed into the row dict."""
        headers = [
            "full_name",
            "email",
            "phone",
            "date_of_birth",
            "gender",
            "address",
            "department_name",
            "position_name",
            "start_date",
            "id_number",
            "tax_code",
            "contract_type",
        ]
        rows = [
            [
                "Nguyen Van A",
                "a@example.com",
                "0901234567",
                "1990-05-15",
                "Male",
                "123 Street, HCMC",
                "Engineering",
                "Developer",
                "2024-01-01",
                "012345678901",
                "1234567890",
                "full_time",
            ]
        ]
        file_bytes = _create_excel(headers, rows)

        parsed, errors = parse_excel(file_bytes)

        assert len(parsed) == 1
        assert len(errors) == 0
        assert parsed[0]["full_name"] == "Nguyen Van A"
        assert parsed[0]["department_name"] == "Engineering"
        assert parsed[0]["contract_type"] == "full_time"

    def test_optional_fields_can_be_empty(self):
        """Optional fields can be None/empty without causing errors."""
        headers = ["full_name", "email", "phone", "gender"]
        rows = [["Nguyen Van A", "a@example.com", None, None]]
        file_bytes = _create_excel(headers, rows)

        parsed, errors = parse_excel(file_bytes)

        assert len(parsed) == 1
        assert len(errors) == 0
        assert parsed[0]["phone"] is None
        assert parsed[0]["gender"] is None


class TestParseExcelDateFormats:
    """Tests for date format handling."""

    def test_date_yyyy_mm_dd_string(self):
        """YYYY-MM-DD string format should be parsed correctly."""
        headers = ["full_name", "email", "date_of_birth"]
        rows = [["Nguyen Van A", "a@example.com", "1990-05-15"]]
        file_bytes = _create_excel(headers, rows)

        parsed, errors = parse_excel(file_bytes)

        assert len(parsed) == 1
        assert len(errors) == 0
        assert parsed[0]["date_of_birth"] == date(1990, 5, 15)

    def test_date_dd_mm_yyyy_string(self):
        """DD/MM/YYYY string format should be parsed correctly."""
        headers = ["full_name", "email", "date_of_birth"]
        rows = [["Nguyen Van A", "a@example.com", "15/05/1990"]]
        file_bytes = _create_excel(headers, rows)

        parsed, errors = parse_excel(file_bytes)

        assert len(parsed) == 1
        assert len(errors) == 0
        assert parsed[0]["date_of_birth"] == date(1990, 5, 15)

    def test_native_datetime_object(self):
        """Native Excel datetime objects should be converted to date."""
        headers = ["full_name", "email", "date_of_birth"]
        # openpyxl stores dates as datetime objects when written as dates
        rows = [["Nguyen Van A", "a@example.com", datetime(1990, 5, 15)]]
        file_bytes = _create_excel(headers, rows)

        parsed, errors = parse_excel(file_bytes)

        assert len(parsed) == 1
        assert len(errors) == 0
        assert parsed[0]["date_of_birth"] == date(1990, 5, 15)

    def test_start_date_parsing(self):
        """start_date field should also support date parsing."""
        headers = ["full_name", "email", "start_date"]
        rows = [["Nguyen Van A", "a@example.com", "01/06/2024"]]
        file_bytes = _create_excel(headers, rows)

        parsed, errors = parse_excel(file_bytes)

        assert len(parsed) == 1
        assert len(errors) == 0
        assert parsed[0]["start_date"] == date(2024, 6, 1)

    def test_invalid_date_format_produces_error(self):
        """An unrecognized date format should produce an error."""
        headers = ["full_name", "email", "date_of_birth"]
        rows = [["Nguyen Van A", "a@example.com", "May 15, 1990"]]
        file_bytes = _create_excel(headers, rows)

        parsed, errors = parse_excel(file_bytes)

        assert len(parsed) == 0
        assert len(errors) == 1
        assert errors[0]["row"] == 2
        assert "Invalid date format" in errors[0]["message"]


class TestParseExcelValidation:
    """Tests for field validation and error reporting."""

    def test_missing_full_name(self):
        """A row missing full_name should produce an error."""
        headers = ["full_name", "email"]
        rows = [[None, "a@example.com"]]
        file_bytes = _create_excel(headers, rows)

        parsed, errors = parse_excel(file_bytes)

        assert len(parsed) == 0
        assert len(errors) == 1
        assert errors[0]["row"] == 2
        assert "full_name" in errors[0]["message"]

    def test_missing_email(self):
        """A row missing email should produce an error."""
        headers = ["full_name", "email"]
        rows = [["Nguyen Van A", None]]
        file_bytes = _create_excel(headers, rows)

        parsed, errors = parse_excel(file_bytes)

        assert len(parsed) == 0
        assert len(errors) == 1
        assert errors[0]["row"] == 2
        assert "email" in errors[0]["message"]

    def test_empty_string_required_field(self):
        """An empty string for a required field should produce an error."""
        headers = ["full_name", "email"]
        rows = [["", "a@example.com"]]
        file_bytes = _create_excel(headers, rows)

        parsed, errors = parse_excel(file_bytes)

        assert len(parsed) == 0
        assert len(errors) == 1
        assert "full_name" in errors[0]["message"]

    def test_invalid_email_format(self):
        """An invalid email format should produce an error."""
        headers = ["full_name", "email"]
        rows = [["Nguyen Van A", "not-an-email"]]
        file_bytes = _create_excel(headers, rows)

        parsed, errors = parse_excel(file_bytes)

        assert len(parsed) == 0
        assert len(errors) == 1
        assert "Invalid email format" in errors[0]["message"]

    def test_multiple_errors_per_row(self):
        """A row with multiple issues should report all errors."""
        headers = ["full_name", "email", "date_of_birth"]
        rows = [[None, "bad-email", "invalid-date"]]
        file_bytes = _create_excel(headers, rows)

        parsed, errors = parse_excel(file_bytes)

        assert len(parsed) == 0
        # Should have errors for: missing full_name, invalid email, invalid date
        assert len(errors) >= 2
        row_numbers = {e["row"] for e in errors}
        assert row_numbers == {2}

    def test_mix_of_valid_and_invalid_rows(self):
        """Valid rows should be parsed even when other rows have errors."""
        headers = ["full_name", "email"]
        rows = [
            ["Nguyen Van A", "a@example.com"],  # valid
            [None, "b@example.com"],  # invalid - missing name
            ["Le Van C", "c@example.com"],  # valid
        ]
        file_bytes = _create_excel(headers, rows)

        parsed, errors = parse_excel(file_bytes)

        assert len(parsed) == 2
        assert len(errors) == 1
        assert errors[0]["row"] == 3  # row 3 in Excel (1-indexed, header is row 1)

    def test_empty_rows_are_skipped(self):
        """Completely empty rows should be silently skipped."""
        headers = ["full_name", "email"]
        rows = [
            ["Nguyen Van A", "a@example.com"],
            [None, None],  # empty row
            ["Le Van C", "c@example.com"],
        ]
        file_bytes = _create_excel(headers, rows)

        parsed, errors = parse_excel(file_bytes)

        assert len(parsed) == 2
        assert len(errors) == 0


class TestParseExcelReadFailures:
    """Tests for file-level read failures, classified by exception type.

    Each case builds a real broken input (no mocking of ``load_workbook``)
    and asserts the returned message is Vietnamese, actionable, and never
    contains library/exception internals.
    """

    _LIBRARY_TERMS = ("BadZipFile", "zip file", "Content_Types", "KeyError", "Traceback")

    def test_empty_bytes_reports_wrong_format(self):
        """Empty bytes are not a valid zip; classified as wrong format."""
        parsed, errors = parse_excel(b"")

        assert parsed == []
        assert len(errors) == 1
        assert errors[0]["row"] == 0
        assert errors[0]["message"] == _MSG_WRONG_FORMAT
        for term in self._LIBRARY_TERMS:
            assert term not in errors[0]["message"]

    def test_plain_text_reports_wrong_format(self):
        """Plain text bytes are not a valid zip; classified as wrong format."""
        parsed, errors = parse_excel(b"just some plain text, not an excel file at all")

        assert parsed == []
        assert len(errors) == 1
        assert errors[0]["message"] == _MSG_WRONG_FORMAT
        for term in self._LIBRARY_TERMS:
            assert term not in errors[0]["message"]

    def test_truncated_xlsx_reports_wrong_format(self):
        """A real .xlsx cut off mid-file breaks the zip container itself."""
        headers = ["full_name", "email"]
        rows = [["Nguyen Van A", "a@example.com"]]
        full_bytes = _create_excel(headers, rows)
        truncated = full_bytes[: len(full_bytes) // 2]

        parsed, errors = parse_excel(truncated)

        assert parsed == []
        assert len(errors) == 1
        assert errors[0]["message"] == _MSG_WRONG_FORMAT
        for term in self._LIBRARY_TERMS:
            assert term not in errors[0]["message"]

    def test_valid_zip_but_not_xlsx_reports_corrupted_file(self):
        """A valid zip missing the OOXML parts is a structurally broken .xlsx."""
        buffer = BytesIO()
        with zipfile.ZipFile(buffer, "w") as zf:
            zf.writestr("hello.txt", "not an excel file")
        file_bytes = buffer.getvalue()

        parsed, errors = parse_excel(file_bytes)

        assert parsed == []
        assert len(errors) == 1
        assert errors[0]["message"] == _MSG_CORRUPTED_FILE
        for term in self._LIBRARY_TERMS:
            assert term not in errors[0]["message"]

    def test_cell_type_mismatch_does_not_leak_cell_content(self):
        """A numeric-typed cell holding unparseable text raises a ValueError
        from openpyxl that echoes the raw cell content in str(e). This must
        never reach the returned message.
        """
        content_types = (
            b'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            b'<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
            b'<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>'  # noqa: E501
            b'<Default Extension="xml" ContentType="application/xml"/>'
            b'<Override PartName="/xl/workbook.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/>'  # noqa: E501
            b'<Override PartName="/xl/worksheets/sheet1.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>'  # noqa: E501
            b"</Types>"
        )
        rels = (
            b'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            b'<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
            b'<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="xl/workbook.xml"/>'  # noqa: E501
            b"</Relationships>"
        )
        workbook_xml = (
            b'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            b'<workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
            b'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">'
            b'<sheets><sheet name="Sheet1" sheetId="1" r:id="rId1"/></sheets>'
            b"</workbook>"
        )
        workbook_rels = (
            b'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            b'<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
            b'<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" Target="worksheets/sheet1.xml"/>'  # noqa: E501
            b"</Relationships>"
        )
        sheet1 = (
            b'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            b'<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
            b"<sheetData>"
            b'<row r="1"><c r="A1" t="n"><v>SECRET_TAX_CODE_998877</v></c></row>'
            b"</sheetData>"
            b"</worksheet>"
        )

        buffer = BytesIO()
        with zipfile.ZipFile(buffer, "w") as zf:
            zf.writestr("[Content_Types].xml", content_types)
            zf.writestr("_rels/.rels", rels)
            zf.writestr("xl/workbook.xml", workbook_xml)
            zf.writestr("xl/_rels/workbook.xml.rels", workbook_rels)
            zf.writestr("xl/worksheets/sheet1.xml", sheet1)
        file_bytes = buffer.getvalue()

        parsed, errors = parse_excel(file_bytes)

        assert parsed == []
        assert len(errors) == 1
        assert "SECRET_TAX_CODE_998877" not in errors[0]["message"]
        assert errors[0]["message"] == _MSG_CORRUPTED_FILE

    def test_invalid_excel_file_logs_server_side(self, caplog: pytest.LogCaptureFixture):
        """A read failure must leave a server-side trace, not just the returned error.

        Previously the exception only reached the caller as a string in the
        returned ``errors`` list — nothing was recorded for operators.
        """
        logger_name = "src.modules.employee.infrastructure.excel_parser"
        with caplog.at_level(logging.ERROR, logger=logger_name):
            parsed, errors = parse_excel(b"garbage bytes string")

        assert parsed == []
        matching = [r for r in caplog.records if r.name == logger_name]
        assert len(matching) == 1
        assert matching[0].levelno == logging.ERROR
        assert matching[0].exc_info is not None


class TestParseExcelEdgeCases:
    """Tests for edge cases."""

    def test_empty_file(self):
        """An empty workbook should return empty results."""
        wb = Workbook()
        buffer = BytesIO()
        wb.save(buffer)
        file_bytes = buffer.getvalue()

        parsed, errors = parse_excel(file_bytes)

        assert parsed == []
        assert errors == []

    def test_header_only_file(self):
        """A file with only headers and no data rows should return empty."""
        headers = ["full_name", "email"]
        file_bytes = _create_excel(headers, [])

        parsed, errors = parse_excel(file_bytes)

        assert parsed == []
        assert errors == []

    def test_headers_with_spaces_and_mixed_case(self):
        """Headers with spaces and mixed case should be normalized."""
        headers = ["Full Name", "Email", "Date Of Birth"]
        rows = [["Nguyen Van A", "a@example.com", "1990-05-15"]]
        file_bytes = _create_excel(headers, rows)

        parsed, errors = parse_excel(file_bytes)

        assert len(parsed) == 1
        assert parsed[0]["full_name"] == "Nguyen Van A"
        assert parsed[0]["date_of_birth"] == date(1990, 5, 15)

    def test_unknown_columns_are_ignored(self):
        """Columns not in KNOWN_FIELDS should be silently ignored."""
        headers = ["full_name", "email", "unknown_column", "another_one"]
        rows = [["Nguyen Van A", "a@example.com", "value1", "value2"]]
        file_bytes = _create_excel(headers, rows)

        parsed, errors = parse_excel(file_bytes)

        assert len(parsed) == 1
        assert "unknown_column" not in parsed[0]
        assert "another_one" not in parsed[0]

    def test_whitespace_in_values_is_stripped(self):
        """Leading/trailing whitespace in string values should be stripped."""
        headers = ["full_name", "email"]
        rows = [["  Nguyen Van A  ", "  a@example.com  "]]
        file_bytes = _create_excel(headers, rows)

        parsed, errors = parse_excel(file_bytes)

        assert len(parsed) == 1
        assert parsed[0]["full_name"] == "Nguyen Van A"
        assert parsed[0]["email"] == "a@example.com"
